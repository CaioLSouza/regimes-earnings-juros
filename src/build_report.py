"""Monta o workbook com as tabelas e graficos do report de regimes da XP,
aplicando a grade earnings x juros. Nao altera nenhum arquivo de entrada."""
import json, math, csv, datetime as dt
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties

import os
PUBLIC = os.environ.get("PUBLIC_BUILD") == "1"
ROOT = os.environ.get("XP_ROOT", "C:/Users/Caio/Documents/Documentos/XP")
AN = f"{ROOT}/tmp/earnings_regimes_analysis"
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = f"{REPO}/data"
OUT = (f"{ROOT}/regimes-earnings-juros/workbook/regimes_earnings_juros_report.xlsx" if PUBLIC
       else f"{ROOT}/outputs/01a06328-b1e5-7e72-a850-3d4674c9b127/regimes_earnings_juros_report.xlsx")

RATE_OVERRIDE = "2025-05-07"
LAST_IPCA = "2026-07"
W = 63
CONFIRM = 2
ORDER = ["EU_JD", "EU_JU", "ED_JD", "ED_JU"]
LABEL = {"EU_JD": "Earnings Up / Juros Down", "EU_JU": "Earnings Up / Juros Up",
         "ED_JD": "Earnings Down / Juros Down", "ED_JU": "Earnings Down / Juros Up"}
COLOR = {"EU_JD": "88ADD8", "EU_JU": "F6B500", "ED_JD": "778078", "ED_JU": "203149"}

# ------------------------------------------------------------------ entradas
wb_eps = openpyxl.load_workbook(f"{ROOT}/Ibovespa Best EPS.xlsx", read_only=True, data_only=True)
eps_rows = sorted((r[0].date().isoformat(), float(r[1]))
                  for r in wb_eps["Sheet1"].iter_rows(min_row=2, values_only=True)
                  if r[0] is not None and r[1] is not None and float(r[1]) > 0)
wb_eps.close()

ibov = sorted((r["data"], float(r["ibov_close"]))
              for r in csv.DictReader(open(f"{AN}/ibov_daily_long.csv", newline="")))

rate_state = {}
for line in open(f"{ROOT}/swap_pre_di_regimes.csv", encoding="utf-8-sig").read().strip().split("\n")[1:]:
    f = line.strip().split(";")
    rate_state[f[0]] = 1 if f[4] == "Hike" else 0 if f[4] == "Cut" else None

swap = {}
for line in open(f"{ROOT}/swap_pre_di_360.csv", encoding="utf-8-sig").read().strip().split("\n")[1:]:
    f = line.strip().split(";")
    try:
        swap[f[0]] = float(f[1].replace(",", "."))
    except (ValueError, IndexError):
        pass

ipca = {}
for fn in ["ipca_2006_2015.json", "ipca_2016_2025.json", "ipca_2026.json"]:
    for r in json.load(open(f"{AN}/ipca/{fn}")):
        d, m, y = r["data"].split("/")
        ipca[f"{y}-{m}"] = float(r["valor"]) / 100

cdi = {}
for r in json.load(open(f"{DATA}/cdi_2006_2013.json")):
    d, m, y = r["data"].split("/")
    cdi[f"{y}-{m}-{d}"] = float(r["valor"]) / 100
for r in csv.DictReader(open(f"{ROOT}/assimetry_score/outputs/corporate/in_sample_cdi_strategy_search/cdi_sgs12.csv", newline="")):
    cdi[r["data"]] = float(r["cdi_daily_return"])

usd = {}
for fn in ["2006_2010", "2011_2015", "2016_2020", "2021_2025", "2026"]:
    for line in open(f"{ROOT}/.codex_eps_analysis/macro/bcb_usdbrl_{fn}.csv", encoding="utf-8-sig").read().strip().split("\n")[1:]:
        f = [x.strip('"') for x in line.strip().split(";")]
        d, m, y = f[0].split("/")
        usd[f"{y}-{m}-{d}"] = float(f[1].replace(",", "."))

# ------------------------------------------------------------ painel e regra
base = []
p = -1
for d, price in ibov:
    while p + 1 < len(eps_rows) and eps_rows[p + 1][0] <= d:
        p += 1
    if p < 0:
        continue
    base.append(dict(date=d, price=price, eps=eps_rows[p][1]))
N = len(base)


def slope(vals):
    n = len(vals)
    xm = (n - 1) / 2
    ym = sum(vals) / n
    return sum((i - xm) * (v - ym) for i, v in enumerate(vals)) / sum((i - xm) ** 2 for i in range(n))


logeps = [math.log(r["eps"]) for r in base]
sl = [None] * N
for i in range(W - 1, N):
    sl[i] = slope(logeps[i - W + 1:i + 1])

month_ends = [(i, 1 if sl[i] > 0 else 0) for i in range(N)
              if sl[i] is not None and (i + 1 == N or base[i]["date"][:7] != base[i + 1]["date"][:7])]
eff, state, pend, pc = {}, None, None, 0
for i, s in month_ends:
    if state is None:
        state = s
        if i + 1 < N:
            eff[i + 1] = s
        continue
    if s == state:
        pend, pc = None, 0
        continue
    pc = pc + 1 if pend == s else 1
    pend = s
    if pc >= CONFIRM and i + 1 < N:
        state = s
        eff[i + 1] = s
        pend, pc = None, 0
earn = [None] * N
act = None
for i in range(N):
    if i in eff:
        act = eff[i]
    earn[i] = act

# as-of para usdbrl e swap
def asof(src, dates):
    keys = sorted(src)
    out, j = {}, -1
    for d in dates:
        while j + 1 < len(keys) and keys[j + 1] <= d:
            j += 1
        out[d] = src[keys[j]] if j >= 0 else None
    return out


all_dates = [r["date"] for r in base]
usd_a = asof(usd, all_dates)
swap_a = asof(swap, all_dates)

panel = []
for i in range(N):
    d = base[i]["date"]
    rt = 0 if d >= RATE_OVERRIDE else rate_state.get(d)
    if earn[i] is None or rt is None:
        continue
    panel.append(dict(date=d, price=base[i]["price"], eps=base[i]["eps"], lneps=logeps[i],
                      slope=sl[i], e=earn[i], r=rt,
                      key=f"{'EU' if earn[i] else 'ED'}_{'JU' if rt else 'JD'}",
                      cdi=cdi.get(d), usd=usd_a[d], swap=swap_a[d]))
n = len(panel)
for k, row in enumerate(panel):
    row["mes"] = row["date"][:7]
    row["tem_ret"] = 1 if k < n - 1 and panel[k + 1]["price"] and row["cdi"] is not None else 0
    row["tem_ipca"] = 1 if row["tem_ret"] and row["mes"] <= LAST_IPCA and row["mes"] in ipca else 0
print("painel:", n, panel[0]["date"], "->", panel[-1]["date"],
      "| intervalos:", sum(r["tem_ret"] for r in panel), "| com ipca:", sum(r["tem_ipca"] for r in panel))

meses = sorted({r["mes"] for r in panel if r["tem_ipca"]})
nint = defaultdict(int)
for r in panel:
    if r["tem_ipca"]:
        nint[r["mes"]] += 1

# episodios combinados
eps_ev = []
s = 0
for i in range(1, n + 1):
    if i == n or panel[i]["key"] != panel[s]["key"]:
        cur = i == n
        end = panel[i - 1]["date"] if cur else panel[i]["date"]
        eps_ev.append(dict(key=panel[s]["key"], start=panel[s]["date"], end=end,
                           end_excl=(dt.date.fromisoformat(panel[-1]["date"]) + dt.timedelta(days=1)).isoformat() if cur else panel[i]["date"],
                           last_td=panel[i - 1]["date"], srow=s + 2, nrows=i - s, cur=cur,
                           dias=(dt.date.fromisoformat(end) - dt.date.fromisoformat(panel[s]["date"])).days))
        s = i
print("episodios combinados:", len(eps_ev))

# ------------------------------------------------------------------ workbook
wb = openpyxl.Workbook()
ARIAL = "Arial"
H1 = Font(name=ARIAL, size=14, bold=True, color="203149")
H2 = Font(name=ARIAL, size=11, bold=True, color="203149")
TH = Font(name=ARIAL, size=9, bold=True, color="FFFFFF")
TXT = Font(name=ARIAL, size=10)
SMALL = Font(name=ARIAL, size=9, color="5F6973")
NAVY = PatternFill("solid", fgColor="203149")
THIN = Side(style="thin", color="D8DCE1")
BOX = Border(bottom=THIN)
PCT = "0.0%"
PCT2 = "0.00%"
NUM0 = "#,##0"
NUM1 = "#,##0.0"
DATE = "dd/mm/yyyy"


def head(ws, row, cols, widths=None):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        cell.font = TH
        cell.fill = NAVY
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w


def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = H1
    if sub:
        ws["A2"] = sub
        ws["A2"].font = SMALL


# ---------------------------------------------------------------- Painel
ws = wb.active
ws.title = "Painel_diario"
PD = "Painel_diario"
cols = ["Data", "Ibovespa",
        "BPA 12m fwd (removido)" if PUBLIC else "BPA 12m fwd",
        "ln(BPA) (removido)" if PUBLIC else "ln(BPA)", "Inclinacao 63d (x1000)", "Earnings", "Juros",
        "Quadrante", "Mes", "Tem retorno", "Tem IPCA", "log ret nominal", "log inflacao dia",
        "log ret real", "CDI ret dia", "log CDI", "log ret USDBRL", "USDBRL", "Swap 360d (%)",
        "EU_JD", "EU_JU", "ED_JD", "ED_JU", "Swap Juros Up", "Swap Juros Down"]
head(ws, 1, cols, [11, 11, 12, 10, 14, 11, 10, 11, 9, 10, 9, 12, 12, 11, 10, 10, 12, 9, 12, 10, 10, 10, 10, 12, 13])
ws.freeze_panes = "B2"
for k, r in enumerate(panel):
    i = k + 2
    ws.cell(i, 1, dt.date.fromisoformat(r["date"])).number_format = DATE
    ws.cell(i, 2, round(r["price"], 2)).number_format = NUM0
    ws.cell(i, 3, None if PUBLIC else round(r["eps"], 4)).number_format = NUM1
    ws.cell(i, 4, None if PUBLIC else f"=LN(C{i})").number_format = "0.0000"
    ws.cell(i, 5, round(r["slope"] * 1000, 4)).number_format = "0.000"
    ws.cell(i, 6, "Up" if r["e"] else "Down")
    ws.cell(i, 7, "Up" if r["r"] else "Down")
    ws.cell(i, 8, r["key"])
    ws.cell(i, 9, r["mes"])
    ws.cell(i, 10, r["tem_ret"])
    ws.cell(i, 11, r["tem_ipca"])
    ws.cell(i, 12, f'=IF($J{i}=1,LN(B{i+1}/B{i}),"")').number_format = "0.000000"
    ws.cell(i, 13, f'=IF($K{i}=1,INDEX(Mensal!$D$2:$D${len(meses)+1},MATCH($I{i},Mensal!$A$2:$A${len(meses)+1},0)),"")').number_format = "0.000000"
    ws.cell(i, 14, f'=IF($K{i}=1,L{i}-M{i},"")').number_format = "0.000000"
    ws.cell(i, 15, r["cdi"]).number_format = "0.000000"
    ws.cell(i, 16, f'=IF($J{i}=1,LN(1+O{i}),"")').number_format = "0.000000"
    ws.cell(i, 17, f'=IF($J{i}=1,LN(R{i+1}/R{i}),"")').number_format = "0.000000"
    ws.cell(i, 18, r["usd"]).number_format = "0.0000"
    ws.cell(i, 19, round(r["swap"], 4) if r["swap"] else None).number_format = "0.00"
    for j, key in enumerate(["EU_JD", "EU_JU", "ED_JD", "ED_JU"]):
        ws.cell(i, 20 + j, round(r["price"], 2) if r["key"] == key else None).number_format = NUM0
    ws.cell(i, 24, round(r["swap"], 4) if (r["swap"] and r["r"] == 1) else None).number_format = "0.00"
    ws.cell(i, 25, round(r["swap"], 4) if (r["swap"] and r["r"] == 0) else None).number_format = "0.00"
LAST = n + 1
DT_ = f"{PD}!$A$2:$A${LAST}"
Q_ = f"{PD}!$H$2:$H${LAST}"
TR_ = f"{PD}!$J$2:$J${LAST}"
TI_ = f"{PD}!$K$2:$K${LAST}"
LN_ = f"{PD}!$L$2:$L${LAST}"
LR_ = f"{PD}!$N$2:$N${LAST}"
LC_ = f"{PD}!$P$2:$P${LAST}"
LU_ = f"{PD}!$Q$2:$Q${LAST}"

# ---------------------------------------------------------------- Mensal
wm = wb.create_sheet("Mensal")
title(wm, "IPCA mensal realizado e alocacao por intervalo de pregao",
      "Fonte: BCB SGS 433 (arquivos JSON em tmp/earnings_regimes_analysis/ipca). log diario = LN(1+IPCA)/intervalos do mes.")
head(wm, 4, ["Mes", "IPCA no mes", "Intervalos de pregao", "log inflacao por intervalo"], [12, 14, 20, 24])
for k, m in enumerate(meses):
    i = k + 5
    wm.cell(i, 1, m)
    wm.cell(i, 2, ipca[m]).number_format = PCT2
    wm.cell(i, 3, nint[m])
    wm.cell(i, 4, f"=LN(1+B{i})/C{i}").number_format = "0.000000"
# a formula do painel aponta para A2/D2: reposiciona cabecalho para linha 1
wm.delete_rows(1, 4)
wm.insert_rows(1)
head(wm, 1, ["Mes", "IPCA no mes", "Intervalos de pregao", "log inflacao por intervalo"], [12, 14, 20, 24])
for k in range(len(meses)):
    i = k + 2
    wm.cell(i, 4).value = f"=LN(1+B{i})/C{i}"
wm.cell(len(meses) + 3, 1, "Fonte: BCB SGS 433. IPCA realizado ate jul/2026; agosto/2026 fica fora do retorno real.").font = SMALL

# ---------------------------------------------------------------- Episodios
we = wb.create_sheet("Episodios")
title(we, "Episodios combinados de earnings e juros",
      "Um episodio comeca no primeiro pregao em que o quadrante vigora e termina no primeiro pregao do quadrante seguinte. Metricas calculadas por formula sobre o Painel_diario.")
hdr = ["#", "Quadrante", "Descricao", "Earnings", "Juros", "Inicio", "Fim", "Fim exclusivo", "Ultimo pregao",
       "Dias corridos", "Pregoes", "Intervalos", "Interv. c/ IPCA", "Ibov total nominal", "Ibov anual nominal",
       "Ibov total real", "Ibov anual real", "CDI anual", "Ibov anual sobre CDI", "USDBRL anual", "Swap 360d (bps)", "Corrente"]
head(we, 4, hdr, [5, 9, 26, 10, 9, 11, 11, 12, 12, 11, 9, 10, 13, 15, 15, 14, 14, 11, 17, 13, 14, 10])
we.freeze_panes = "A5"
E0 = 5
for k, ev in enumerate(eps_ev):
    i = E0 + k
    cnt = f'COUNTIFS({DT_},">="&$F{i},{DT_},"<"&$H{i},{TR_},1)'
    cnti = f'COUNTIFS({DT_},">="&$F{i},{DT_},"<"&$H{i},{TI_},1)'
    sn = f'SUMIFS({LN_},{DT_},">="&$F{i},{DT_},"<"&$H{i})'
    sr = f'SUMIFS({LR_},{DT_},">="&$F{i},{DT_},"<"&$H{i})'
    sc = f'SUMIFS({LC_},{DT_},">="&$F{i},{DT_},"<"&$H{i})'
    su = f'SUMIFS({LU_},{DT_},">="&$F{i},{DT_},"<"&$H{i})'
    scr = f'SUMIFS({LC_},{DT_},">="&$F{i},{DT_},"<"&$H{i},{TI_},1)'
    snr = f'SUMIFS({LN_},{DT_},">="&$F{i},{DT_},"<"&$H{i},{TI_},1)'
    we.cell(i, 1, k + 1)
    we.cell(i, 2, ev["key"])
    we.cell(i, 3, LABEL[ev["key"]])
    we.cell(i, 4, "Up" if ev["key"][1] == "U" else "Down")
    we.cell(i, 5, "Up" if ev["key"][-2:] == "JU" else "Down")
    we.cell(i, 6, dt.date.fromisoformat(ev["start"])).number_format = DATE
    we.cell(i, 7, dt.date.fromisoformat(ev["end"])).number_format = DATE
    we.cell(i, 8, dt.date.fromisoformat(ev["end_excl"])).number_format = DATE
    we.cell(i, 9, dt.date.fromisoformat(ev["last_td"])).number_format = DATE
    we.cell(i, 10, f"=G{i}-F{i}").number_format = NUM0
    we.cell(i, 11, f'=COUNTIFS({DT_},">="&$F{i},{DT_},"<"&$H{i})').number_format = NUM0
    we.cell(i, 12, f"={cnt}").number_format = NUM0
    we.cell(i, 13, f"={cnti}").number_format = NUM0
    we.cell(i, 14, f"=EXP({sn})-1").number_format = PCT
    we.cell(i, 15, f"=IF(L{i}=0,\"\",EXP({sn}*252/L{i})-1)").number_format = PCT
    we.cell(i, 16, f"=EXP({sr})-1").number_format = PCT
    we.cell(i, 17, f"=IF(M{i}=0,\"\",EXP({sr}*252/M{i})-1)").number_format = PCT
    we.cell(i, 18, f"=IF(L{i}=0,\"\",EXP({sc}*252/L{i})-1)").number_format = PCT
    we.cell(i, 19, f"=IF(L{i}=0,\"\",EXP(({sn}-{sc})*252/L{i})-1)").number_format = PCT
    we.cell(i, 20, f"=IF(L{i}=0,\"\",EXP({su}*252/L{i})-1)").number_format = PCT
    we.cell(i, 21, f'=(INDEX({PD}!$S$2:$S${LAST},MATCH($I{i},{DT_},0))-INDEX({PD}!$S$2:$S${LAST},MATCH($F{i},{DT_},0)))*100').number_format = NUM0
    we.cell(i, 22, "Sim" if ev["cur"] else "")
    for j in range(1, 23):
        we.cell(i, j).font = TXT
        we.cell(i, j).border = BOX
EN = E0 + len(eps_ev) - 1
we.cell(EN + 2, 1, "Retorno anualizado usa capitalizacao log com 252 pregoes. O real desconta o IPCA do mes alocado por intervalo. USDBRL positivo = real se desvalorizando.").font = SMALL
we.cell(EN + 3, 1, "Swap 360d (bps) = variacao do nivel da taxa entre o primeiro e o ultimo pregao do episodio.").font = SMALL
we.cell(EN + 4, 1, "O episodio corrente termina em 12/08/2026 e nao tem IPCA de agosto, por isso os intervalos com IPCA sao menores que os intervalos totais.").font = SMALL

# ---------------------------------------------------------------- Resumo
wr = wb.create_sheet("Resumo_regimes", 1)
title(wr, "Ibovespa nos regimes de earnings e juros", "Amostra efetiva de 03/04/2006 a 12/08/2026. Todas as celulas abaixo sao formulas sobre Painel_diario e Episodios.")
head(wr, 4, ["Quadrante", "Descricao", "% do tempo", "Dias corridos", "Episodios", "Duracao media (dias)",
             "Duracao mediana (dias)", "Ibov anual nominal", "Ibov anual real", "Ibov anual sobre CDI",
             "Dias positivos", "Episodios positivos", "Taxa de acerto"],
     [10, 26, 11, 12, 10, 14, 15, 14, 13, 15, 11, 13, 12])
R0 = 5
for k, key in enumerate(ORDER):
    i = R0 + k
    cnt = f'COUNTIFS({Q_},$A{i},{TR_},1)'
    cnti = f'COUNTIFS({Q_},$A{i},{TI_},1)'
    wr.cell(i, 1, key).font = TXT
    wr.cell(i, 2, LABEL[key]).font = TXT
    wr.cell(i, 3, f"=D{i}/SUM($D${R0}:$D${R0+3})").number_format = PCT
    wr.cell(i, 4, f'=SUMIF(Episodios!$B${E0}:$B${EN},$A{i},Episodios!$J${E0}:$J${EN})').number_format = NUM0
    wr.cell(i, 5, f'=COUNTIF(Episodios!$B${E0}:$B${EN},$A{i})').number_format = NUM0
    wr.cell(i, 6, f"=D{i}/E{i}").number_format = NUM0
    wr.cell(i, 8, f'=EXP(SUMIFS({LN_},{Q_},$A{i})*252/{cnt})-1').number_format = PCT
    wr.cell(i, 9, f'=EXP(SUMIFS({LR_},{Q_},$A{i})*252/{cnti})-1').number_format = PCT
    wr.cell(i, 10, f'=EXP((SUMIFS({LN_},{Q_},$A{i})-SUMIFS({LC_},{Q_},$A{i}))*252/{cnt})-1').number_format = PCT
    wr.cell(i, 11, f'=COUNTIFS({Q_},$A{i},{LN_},">0")/{cnt}').number_format = PCT
    wr.cell(i, 12, f'=COUNTIFS(Episodios!$B${E0}:$B${EN},$A{i},Episodios!$P${E0}:$P${EN},">0")').number_format = NUM0
    wr.cell(i, 13, f"=L{i}/E{i}").number_format = PCT
    for j in range(1, 14):
        wr.cell(i, j).font = TXT
        wr.cell(i, j).border = BOX
wr.cell(R0 + 4, 2, "Total / Ibovespa na amostra").font = Font(name=ARIAL, size=10, bold=True)
wr.cell(R0 + 4, 3, f"=SUM(C{R0}:C{R0+3})").number_format = PCT
wr.cell(R0 + 4, 4, f"=SUM(D{R0}:D{R0+3})").number_format = NUM0
wr.cell(R0 + 4, 5, f"=SUM(E{R0}:E{R0+3})").number_format = NUM0
wr.cell(R0 + 4, 8, f'=EXP(SUM({LN_})*252/COUNTIF({TR_},1))-1').number_format = PCT
wr.cell(R0 + 4, 9, f'=EXP(SUM({LR_})*252/COUNTIF({TI_},1))-1').number_format = PCT
wr.cell(R0 + 4, 10, f'=EXP((SUM({LN_})-SUM({LC_}))*252/COUNTIF({TR_},1))-1').number_format = PCT
wr.cell(R0 + 4, 11, f'=COUNTIF({LN_},">0")/COUNTIF({TR_},1)').number_format = PCT
for j in range(1, 14):
    wr.cell(R0 + 4, j).font = Font(name=ARIAL, size=10, bold=True)

# marginais
M0 = R0 + 7
wr.cell(M0 - 1, 1, "Frequencias marginais").font = H2
head(wr, M0, ["Classificacao", "Dias corridos", "% do tempo", "Ibov anual real"], [16, 14, 12, 14])
marg = [("Earnings Up", "F", "Up"), ("Earnings Down", "F", "Down"), ("Juros Up", "G", "Up"), ("Juros Down", "G", "Down")]
for k, (lab, col, val) in enumerate(marg):
    i = M0 + 1 + k
    rng = f"{PD}!${col}$2:${col}${LAST}"
    keys = [q for q in ORDER if (q[1] == "U") == (val == "Up")] if col == "F" else [q for q in ORDER if (q[-2:] == "JU") == (val == "Up")]
    dias = "+".join([f'SUMIF(Episodios!$B${E0}:$B${EN},"{q}",Episodios!$J${E0}:$J${EN})' for q in keys])
    wr.cell(i, 1, lab).font = TXT
    wr.cell(i, 2, f"={dias}").number_format = NUM0
    wr.cell(i, 3, f"=B{i}/SUM($D${R0}:$D${R0+3})").number_format = PCT
    wr.cell(i, 4, f'=EXP(SUMIFS({LR_},{rng},"{val}")*252/COUNTIFS({rng},"{val}",{TI_},1))-1').number_format = PCT
    for j in range(1, 5):
        wr.cell(i, j).border = BOX

S0 = M0 + 7
wr.cell(S0 - 1, 1, "Estado corrente em 12/08/2026").font = H2
cur_e = "Earnings Up" if panel[-1]["e"] else "Earnings Down"
cur_r = "Juros Down" if panel[-1]["r"] == 0 else "Juros Up"
for k, (a, b) in enumerate([("Regime de earnings", f"{cur_e} desde 01/12/2025"),
                            ("Regime de juros", f"{cur_r} desde 07/05/2025"),
                            ("Quadrante combinado", f"{LABEL[panel[-1]['key']]} desde 01/12/2025"),
                            ("Ultima observacao comum", "12/08/2026"),
                            ("Ultimo IPCA realizado", "julho de 2026")]):
    wr.cell(S0 + k, 1, a).font = Font(name=ARIAL, size=10, bold=True)
    wr.cell(S0 + k, 2, b).font = TXT
wr.cell(S0 + 7, 1, "Duracao mediana preenchida na aba Ep_por_quadrante, onde os episodios de cada quadrante ficam em linhas contiguas.").font = SMALL

ch = BarChart()
ch.type = "col"
ch.style = 2
ch.title = "Ibovespa: retorno real anualizado por quadrante"
ch.y_axis.numFmt = "0%"
ch.y_axis.title = "Retorno real anualizado"
data = Reference(wr, min_col=9, min_row=4, max_row=R0 + 3)
cats = Reference(wr, min_col=2, min_row=R0, max_row=R0 + 3)
ch.add_data(data, titles_from_data=True)
ch.set_categories(cats)
ch.width, ch.height = 20, 10
ch.legend = None
try:
    pts = ch.series[0]
    from openpyxl.chart.marker import DataPoint
    pts.data_points = [DataPoint(idx=i, spPr=GraphicalProperties(solidFill=COLOR[k])) for i, k in enumerate(ORDER)]
except Exception as exc:
    print("cor das barras:", exc)
wr.add_chart(ch, "P4")

ch2 = BarChart()
ch2.type = "col"
ch2.title = "Frequencia de cada quadrante (dias corridos)"
ch2.y_axis.numFmt = "0%"
d2 = Reference(wr, min_col=3, min_row=4, max_row=R0 + 3)
ch2.add_data(d2, titles_from_data=True)
ch2.set_categories(cats)
ch2.width, ch2.height = 20, 9
ch2.legend = None
try:
    from openpyxl.chart.marker import DataPoint
    ch2.series[0].data_points = [DataPoint(idx=i, spPr=GraphicalProperties(solidFill=COLOR[k])) for i, k in enumerate(ORDER)]
except Exception as exc:
    print("cor das barras 2:", exc)
wr.add_chart(ch2, "P25")

# ------------------------------------------------------- Ep por quadrante
wq = wb.create_sheet("Ep_por_quadrante")
title(wq, "Episodios por quadrante", "Mesmo formato das tabelas do report da XP: primeiro sinal, fim e duracao, com o desempenho do Ibovespa em cada janela.")
row = 4
qrow = {}
for key in ORDER:
    wq.cell(row, 1, LABEL[key]).font = H2
    head(wq, row + 1, ["Primeiro sinal", "Fim", "Duracao (dias)", "Ibov total nominal", "Ibov anual real", "Ibov anual sobre CDI"],
         [14, 14, 14, 16, 15, 18])
    idxs = [k for k, ev in enumerate(eps_ev) if ev["key"] == key]
    for j, k in enumerate(idxs):
        i = row + 2 + j
        src = E0 + k
        wq.cell(i, 1, f"=Episodios!F{src}").number_format = DATE
        wq.cell(i, 2, f'=IF(Episodios!V{src}="Sim","Atual",Episodios!G{src})').number_format = DATE
        wq.cell(i, 3, f"=Episodios!J{src}").number_format = NUM0
        wq.cell(i, 4, f"=Episodios!N{src}").number_format = PCT
        wq.cell(i, 5, f"=Episodios!Q{src}").number_format = PCT
        wq.cell(i, 6, f"=Episodios!S{src}").number_format = PCT
        for c in range(1, 7):
            wq.cell(i, c).font = TXT
            wq.cell(i, c).border = BOX
    last = row + 1 + len(idxs)
    qrow[key] = (row + 2, last)
    wq.cell(last + 1, 1, "Mediana").font = Font(name=ARIAL, size=10, bold=True)
    wq.cell(last + 1, 3, f"=MEDIAN(C{row+2}:C{last})").number_format = NUM0
    wq.cell(last + 1, 5, f"=MEDIAN(E{row+2}:E{last})").number_format = PCT
    wq.cell(last + 2, 1, "Media ponderada por dias").font = Font(name=ARIAL, size=10, bold=True)
    wq.cell(last + 2, 5, f"=Resumo_regimes!I{R0+ORDER.index(key)}").number_format = PCT
    wq.cell(last + 2, 6, f"=Resumo_regimes!J{R0+ORDER.index(key)}").number_format = PCT
    wq.cell(last + 3, 1, "Taxa de acerto (episodios com retorno real positivo)").font = Font(name=ARIAL, size=10, bold=True)
    wq.cell(last + 3, 5, f"=Resumo_regimes!M{R0+ORDER.index(key)}").number_format = PCT
    row = last + 6
for key in ORDER:
    a, b = qrow[key]
    wr.cell(R0 + ORDER.index(key), 7, f"=MEDIAN(Ep_por_quadrante!C{a}:C{b})").number_format = NUM0

# ------------------------------------------------------------ Transicoes
wt = wb.create_sheet("Transicoes")
title(wt, "Matriz de transicao entre quadrantes", "Com que frequencia, historicamente, cada quadrante foi seguido por cada outro. Equivalente ao diagrama de transicao do report da XP.")
trans = defaultdict(lambda: defaultdict(int))
for a, b in zip(eps_ev, eps_ev[1:]):
    trans[a["key"]][b["key"]] += 1
head(wt, 4, ["De \\ Para"] + [LABEL[k] for k in ORDER] + ["Total de transicoes"], [26, 22, 22, 22, 22, 18])
for k, key in enumerate(ORDER):
    i = 5 + k
    wt.cell(i, 1, LABEL[key]).font = Font(name=ARIAL, size=10, bold=True)
    tot = sum(trans[key].values())
    for j, key2 in enumerate(ORDER):
        c = wt.cell(i, 2 + j, (trans[key][key2] / tot) if tot else 0)
        c.number_format = PCT
        c.font = TXT
        c.border = BOX
        if key == key2:
            c.fill = PatternFill("solid", fgColor="F0F0F0")
    wt.cell(i, 6, tot).font = TXT
wt.cell(10, 1, "Contagens absolutas").font = H2
head(wt, 11, ["De \\ Para"] + [LABEL[k] for k in ORDER] + ["Total"], [26, 22, 22, 22, 22, 18])
for k, key in enumerate(ORDER):
    i = 12 + k
    wt.cell(i, 1, LABEL[key]).font = Font(name=ARIAL, size=10, bold=True)
    for j, key2 in enumerate(ORDER):
        wt.cell(i, 2 + j, trans[key][key2]).font = TXT
    wt.cell(i, 6, f"=SUM(B{i}:E{i})").font = TXT
wt.cell(17, 1, "Transicoes contadas sobre 41 episodios, portanto 40 transicoes. A diagonal e zero por construcao: um episodio so termina quando o quadrante muda.").font = SMALL
wt.cell(18, 1, "Uma mudanca de quadrante pode vir do lado de earnings ou do lado de juros, nunca dos dois no mesmo dia, exceto por coincidencia de datas.").font = SMALL

# --------------------------------------------------------- Classes de ativos
wa = wb.create_sheet("Classes_ativos")
title(wa, "Desempenho de classes de ativos em cada episodio",
      "Formato da tabela do report da XP. Nao ha SMLL, IMA-B nem cestas de fatores nos dados do projeto, entao as linhas disponiveis sao Ibovespa, CDI, USDBRL e o nivel do swap.")
row = 4
for key in ORDER:
    idxs = [k for k, ev in enumerate(eps_ev) if ev["key"] == key]
    wa.cell(row, 1, LABEL[key]).font = H2
    wa.cell(row + 1, 1, "Inicio").font = TH
    wa.cell(row + 1, 1).fill = NAVY
    wa.cell(row + 2, 1, "Fim").font = TH
    wa.cell(row + 2, 1).fill = NAVY
    wa.cell(row + 3, 1, "Duracao (dias)").font = TH
    wa.cell(row + 3, 1).fill = NAVY
    wa.column_dimensions["A"].width = 30
    for j, k in enumerate(idxs):
        c = 2 + j
        src = E0 + k
        wa.column_dimensions[get_column_letter(c)].width = 11
        wa.cell(row + 1, c, f"=Episodios!F{src}").number_format = "mmm/yy"
        wa.cell(row + 2, c, f'=IF(Episodios!V{src}="Sim","Atual",Episodios!G{src})').number_format = "mmm/yy"
        wa.cell(row + 3, c, f"=Episodios!J{src}").number_format = NUM0
    wavg = 2 + len(idxs)
    whit = wavg + 1
    wa.column_dimensions[get_column_letter(wavg)].width = 14
    wa.column_dimensions[get_column_letter(whit)].width = 13
    wa.cell(row + 1, wavg, "Media ponderada").font = TH
    wa.cell(row + 1, wavg).fill = NAVY
    wa.cell(row + 1, whit, "Taxa de acerto").font = TH
    wa.cell(row + 1, whit).fill = NAVY
    lines = [("Ibovespa nominal anualizado", "O", 8), ("Ibovespa real anualizado", "Q", 9),
             ("Ibovespa sobre o CDI, anualizado", "S", 10), ("CDI anualizado", "R", None),
             ("USDBRL anualizado", "T", None), ("Swap 360d, variacao (bps)", "U", None)]
    for li, (lab, col, rescol) in enumerate(lines):
        i = row + 4 + li
        wa.cell(i, 1, lab).font = TXT
        wa.cell(i, 1).border = BOX
        for j, k in enumerate(idxs):
            src = E0 + k
            c = wa.cell(i, 2 + j, f"=Episodios!{col}{src}")
            c.number_format = NUM0 if col == "U" else PCT
            c.font = TXT
            c.border = BOX
        rng = ",".join([f"Episodios!{col}{E0+k}" for k in idxs])
        if rescol:
            wa.cell(i, wavg, f"=Resumo_regimes!{get_column_letter(rescol)}{R0+ORDER.index(key)}").number_format = PCT
        else:
            first = get_column_letter(2)
            lastc = get_column_letter(1 + len(idxs))
            wa.cell(i, wavg, f"=SUMPRODUCT({first}{i}:{lastc}{i},${first}${row+3}:${lastc}${row+3})/SUM(${first}${row+3}:${lastc}${row+3})").number_format = NUM0 if col == "U" else PCT
        first = get_column_letter(2)
        lastc = get_column_letter(1 + len(idxs))
        wa.cell(i, whit, f'=COUNTIF({first}{i}:{lastc}{i},">0")/COUNT({first}{i}:{lastc}{i})').number_format = PCT
        wa.cell(i, wavg).font = TXT
        wa.cell(i, whit).font = TXT
    row = row + 4 + len(lines) + 3
wa.cell(row, 1, "Media ponderada das linhas do Ibovespa vem do agregado do quadrante em Resumo_regimes; as demais sao media ponderada pela duracao de cada episodio.").font = SMALL
wa.cell(row + 1, 1, "CDI reconstruido do BCB SGS 12, diario, de 2006 a 2026. USDBRL do BCB, PTAX diaria. Nao ha dados de SMLL, IMA-B 5, IMA-B 5+, setores ou fatores no projeto.").font = SMALL

# ------------------------------------------------------ Trajetoria eventos
wtr = wb.create_sheet("Trajetoria_eventos")
cur_key = panel[-1]["key"]
ev_cur = [k for k, ev in enumerate(eps_ev) if ev["key"] == cur_key][-5:]
title(wtr, f"Ibovespa nos ultimos 5 episodios de {LABEL[cur_key]}",
      "Retorno nominal acumulado a partir do primeiro pregao de cada episodio, ate 252 pregoes. Equivale ao grafico de ultimos eventos do report da XP.")
hdrs = ["Pregao"] + [eps_ev[k]["start"] for k in ev_cur]
head(wtr, 4, hdrs, [10] + [14] * len(ev_cur))
wtr.cell(5, 1, "Pregoes do episodio").font = Font(name=ARIAL, size=9, bold=True)
for j, k in enumerate(ev_cur):
    wtr.cell(5, 2 + j, eps_ev[k]["nrows"]).font = Font(name=ARIAL, size=9, bold=True)
for h in range(0, 253):
    i = 6 + h
    wtr.cell(i, 1, h)
    for j, k in enumerate(ev_cur):
        ev = eps_ev[k]
        if h == 0:
            wtr.cell(i, 2 + j, 0).number_format = PCT
        else:
            a = ev["srow"]
            b = ev["srow"] + h - 1
            if b <= LAST:
                wtr.cell(i, 2 + j, f"=EXP(SUM({PD}!$L${a}:$L${b}))-1").number_format = PCT
        wtr.cell(i, 2 + j).font = TXT
lc = LineChart()
lc.title = f"Ibovespa nos ultimos 5 episodios de {LABEL[cur_key]}"
lc.y_axis.numFmt = "0%"
lc.y_axis.title = "Retorno nominal acumulado"
lc.x_axis.title = "Pregoes desde o primeiro sinal"
lc.x_axis.delete = False
lc.y_axis.delete = False
for j, k in enumerate(ev_cur):
    ref = Reference(wtr, min_col=2 + j, min_row=6, max_row=258)
    ser = Series(ref, title=eps_ev[k]["start"])
    ser.smooth = False
    ser.marker = Marker(symbol="none")
    lc.series.append(ser)
lc.set_categories(Reference(wtr, min_col=1, min_row=6, max_row=258))
lc.width, lc.height = 24, 12
wtr.add_chart(lc, "H4")
wtr.cell(261, 1, "Convencao de estudo de evento, igual a do report da XP: a curva segue por 252 pregoes a partir do primeiro sinal, mesmo que o episodio termine antes. A linha 5 mostra quantos pregoes cada episodio durou.").font = SMALL
wtr.cell(262, 1, "Depois desse ponto a curva ja mistura o quadrante seguinte, entao a leitura util e o trecho ate a duracao indicada.").font = SMALL

# ---------------------------------------------------------------- Graficos
wg = wb.create_sheet("Graficos", 2)
title(wg, "Graficos do report", "As series vem do Painel_diario. O Ibovespa aparece em quatro colunas, uma por quadrante, para colorir o nivel do indice conforme o regime vigente.")
lc2 = LineChart()
lc2.title = "Ibovespa ao longo dos regimes de earnings e juros"
lc2.y_axis.title = "Pontos"
lc2.x_axis.title = "Data"
lc2.x_axis.delete = False
lc2.y_axis.delete = False
for j, key in enumerate(["EU_JD", "EU_JU", "ED_JD", "ED_JU"]):
    ref = Reference(ws, min_col=20 + j, min_row=1, max_row=LAST)
    s = Series(ref, title=LABEL[key])
    s.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill=COLOR[key], w=20000))
    s.marker = Marker(symbol="none")
    s.smooth = False
    lc2.series.append(s)
lc2.set_categories(Reference(ws, min_col=1, min_row=2, max_row=LAST))
lc2.width, lc2.height = 32, 13
wg.add_chart(lc2, "A4")

lc3 = LineChart()
lc3.title = "Swap Pre-DI 360 dias e os regimes de juros"
lc3.y_axis.title = "% ao ano"
lc3.x_axis.title = "Data"
lc3.x_axis.delete = False
lc3.y_axis.delete = False
for j, (col, lab, color) in enumerate([(24, "Juros Up (Hike)", "203149"), (25, "Juros Down (Cut)", "88ADD8")]):
    ref = Reference(ws, min_col=col, min_row=1, max_row=LAST)
    s = Series(ref, title=lab)
    s.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill=color, w=20000))
    s.marker = Marker(symbol="none")
    s.smooth = False
    lc3.series.append(s)
lc3.set_categories(Reference(ws, min_col=1, min_row=2, max_row=LAST))
lc3.width, lc3.height = 32, 12
wg.add_chart(lc3, "A32")

if not PUBLIC:
  lc4 = LineChart()
  lc4.title = "BPA esperado 12 meses a frente do Ibovespa"
  lc4.y_axis.title = "Pontos de BPA"
  lc4.x_axis.delete = False
  lc4.y_axis.delete = False
  ref = Reference(ws, min_col=3, min_row=1, max_row=LAST)
  s = Series(ref, title="BPA 12m fwd")
  s.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="203149", w=18000))
  s.marker = Marker(symbol="none")
  s.smooth = False
  lc4.series.append(s)
  lc4.set_categories(Reference(ws, min_col=1, min_row=2, max_row=LAST))
  lc4.width, lc4.height = 32, 11
  wg.add_chart(lc4, "A60")

# -------------------------------------------------------------- Diagnostico
wd = wb.create_sheet("Diagnostico")
title(wd, "Robustez da regra e limites de leitura",
      "Numeros apurados na auditoria independente. Servem para acompanhar as tabelas acima e evitar que os pontos sejam lidos como precisos.")
wd.cell(4, 1, "Sensibilidade: retorno real anualizado por quadrante ao variar janela e confirmacoes").font = H2
head(wd, 5, ["Janela x confirmacoes", "EU_JD", "ED_JD", "EU_JU", "ED_JU", "Spread em Juros Down", "Spread em Juros Up", "Episodios de earnings"],
     [22, 10, 10, 10, 10, 20, 18, 20])
sens = [("42 x 2", 21.5, 13.0, -9.7, -15.6, 8.6, 5.9, 35), ("63 x 1", 22.7, 11.5, -8.9, -17.0, 11.2, 8.1, 47),
        ("63 x 2 (oficial)", 26.8, 8.4, -8.3, -18.3, 18.4, 10.0, 29), ("63 x 3", 17.8, 16.5, -5.4, -27.8, 1.2, 22.3, 19),
        ("84 x 2", 29.2, 8.4, -9.3, -17.9, 20.8, 8.6, 26), ("84 x 3", 15.3, 20.3, -7.4, -22.7, -5.0, 15.4, 18),
        ("126 x 2", 12.7, 23.8, -8.8, -18.2, -11.1, 9.4, 22)]
for k, r in enumerate(sens):
    i = 6 + k
    for j, v in enumerate(r):
        c = wd.cell(i, 1 + j, v)
        c.font = Font(name=ARIAL, size=10, bold=(k == 2))
        c.border = BOX
        if j in (1, 2, 3, 4, 5, 6):
            c.number_format = "0.0"
wd.cell(14, 1, 'A regra oficial usa 63 pregoes e 2 confirmacoes. O spread de earnings dentro de Juros Down vai de -11,1 a +20,8 p.p. na vizinhanca, entao a magnitude nao e um numero preciso.').font = SMALL

wd.cell(17, 1, "Teste de rotacao circular do rotulo de earnings").font = H2
head(wd, 18, ["Estatistica", "Observado (p.p.)", "p bicaudal", "Faixa nula 5-95%"], [34, 16, 12, 20])
rot = [("Spread de earnings dentro de Juros Down", 15.7, 0.322, "-25,3 a +25,3"),
       ("Spread de earnings dentro de Juros Up", 11.5, 0.509, "-25,7 a +27,4"),
       ("Spread de earnings incondicional", 7.7, 0.523, "-18,3 a +21,5"),
       ("Spread de juros incondicional (referencia)", 28.1, 0.016, "-16,2 a +16,9")]
for k, r in enumerate(rot):
    i = 19 + k
    for j, v in enumerate(r):
        c = wd.cell(i, 1 + j, v)
        c.font = TXT
        c.border = BOX
        if j == 1:
            c.number_format = "0.0"
        if j == 2:
            c.number_format = "0.000"
wd.cell(24, 1, "A rotacao preserva a persistencia do regime e sorteia o timing. O rotulo de juros passa no teste; o de earnings nao. Spread aritmetico anualizado.").font = SMALL

wd.cell(27, 1, "Bootstrap por episodio, reamostrando os 41 episodios inteiros").font = H2
head(wd, 28, ["Spread de earnings", "Observado (p.p.)", "IC 90% inferior", "IC 90% superior", "Probabilidade de ser positivo"], [24, 16, 16, 16, 26])
for k, r in enumerate([("Dentro de Juros Down", 15.7, -4.5, 31.2, 0.90), ("Dentro de Juros Up", 11.5, -10.5, 34.0, 0.82)]):
    i = 29 + k
    for j, v in enumerate(r):
        c = wd.cell(i, 1 + j, v)
        c.font = TXT
        c.border = BOX
        if j in (1, 2, 3):
            c.number_format = "0.0"
        if j == 4:
            c.number_format = PCT

wd.cell(33, 1, "Poder preditivo em horizontes de carteira (spread de earnings, p.p.)").font = H2
head(wd, 34, ["Horizonte", "Dentro de Juros Down", "Dentro de Juros Up"], [16, 22, 20])
for k, r in enumerate([("1 mes", -0.6, 10.9), ("3 meses", -9.4, 2.6), ("6 meses", 2.2, 1.9), ("12 meses", 2.9, 3.5)]):
    i = 35 + k
    for j, v in enumerate(r):
        c = wd.cell(i, 1 + j, v)
        c.font = TXT
        c.border = BOX
        if j:
            c.number_format = "0.0"
wd.cell(40, 1, "O spread desaparece e troca de sinal em qualquer horizonte de carteira. A grade descreve o ambiente vigente; nao antecipa o retorno.").font = SMALL

wd.cell(43, 1, "Subamostras: retorno real anualizado por quadrante (%)").font = H2
head(wd, 44, ["Periodo", "EU_JD", "ED_JD", "EU_JU", "ED_JU", "Spread em Juros Down", "Spread em Juros Up"], [16, 10, 10, 10, 10, 20, 18])
for k, r in enumerate([("2006-2015", 40.8, 11.8, -16.0, -20.5, 29.1, 4.5), ("2016-2026", 23.1, 5.2, 1.4, -11.2, 17.9, 12.6),
                       ("2006-2012", 40.8, 17.8, -18.7, -20.4, 23.0, 1.6), ("2013-2019", 20.3, 21.3, 6.1, -21.9, -0.9, 28.1),
                       ("2020-2026", 27.7, -11.0, -3.5, -4.7, 38.7, 1.2)]):
    i = 45 + k
    for j, v in enumerate(r):
        c = wd.cell(i, 1 + j, v)
        c.font = TXT
        c.border = BOX
        if j:
            c.number_format = "0.0"
wd.cell(51, 1, "A ordenacao Juros Down acima de Juros Up vale em todas as subamostras. A de earnings vale em 9 das 10 celulas, com magnitude muito instavel.").font = SMALL

# ------------------------------------------------------------------ Leia-me
wl = wb.create_sheet("Leia-me", 0)
wl.column_dimensions["A"].width = 34
wl.column_dimensions["B"].width = 104
wl["A1"] = "Regimes de earnings e juros do Ibovespa"
wl["A1"].font = Font(name=ARIAL, size=16, bold=True, color="203149")
wl["A2"] = "Tabelas e graficos no formato do Regime Monitor da XP, com a grade earnings x juros"
wl["A2"].font = Font(name=ARIAL, size=11, color="5F6973")
info = [
    ("", ""),
    ("Regra de earnings", "Inclinacao OLS do log do BPA esperado 12 meses a frente sobre os ultimos 63 pregoes. O sinal e observado apenas no ultimo pregao de cada mes. Inclinacao positiva e Earnings Up; nao positiva e Earnings Down."),
    ("Filtro de mudanca", "Uma virada exige dois fechamentos mensais consecutivos do lado oposto e passa a valer no pregao seguinte a segunda confirmacao. Sem media movel, sem winsorizacao, sem banda e sem ajuste pelo Focus."),
    ("Regra de juros", "Classificacao Hike e Cut da serie do swap Pre-DI de 360 dias. Hike e Juros Up, Cut e Juros Down. Juros Down forcado desde 07/05/2025 ate o fim da amostra, em linha com os reports da XP."),
    ("Retornos", "O retorno de fechamento a fechamento de t para t+1 e atribuido ao regime vigente em t. Composicao em log, anualizacao por 252 pregoes."),
    ("Retorno real", "IPCA mensal realizado, com o log da inflacao do mes distribuido igualmente entre os intervalos de pregao daquele mes. Agosto de 2026 fica fora do retorno real por nao haver IPCA realizado."),
    ("Frequencia", "Dias corridos entre datas efetivas de mudanca de regime."),
    ("Amostra", "03/04/2006 a 12/08/2026, 5.036 pregoes classificados, 29 episodios de earnings e 41 episodios combinados."),
    ("", ""),
    ("Abas", ""),
    ("Resumo_regimes", "Quadro-sintese dos quatro quadrantes, frequencias marginais, estado corrente e os dois graficos de barras."),
    ("Graficos", "Ibovespa colorido pelo quadrante vigente, swap Pre-DI por regime de juros e a serie de BPA."),
    ("Episodios", "Os 41 episodios combinados com duracao e desempenho de Ibovespa, CDI, USDBRL e swap."),
    ("Ep_por_quadrante", "Os mesmos episodios separados por quadrante, no formato das tabelas do report, com mediana e taxa de acerto."),
    ("Classes_ativos", "Grade de classes de ativos por episodio, com media ponderada e taxa de acerto."),
    ("Transicoes", "Matriz de transicao entre quadrantes, equivalente ao diagrama do report."),
    ("Trajetoria_eventos", "Retorno acumulado do Ibovespa nos ultimos cinco episodios do quadrante corrente, ate 252 pregoes."),
    ("Painel_diario", "Base diaria completa. Todas as tabelas apontam para ela."),
    ("Mensal", "IPCA por mes e a alocacao do log da inflacao entre os intervalos de pregao."),
    ("Diagnostico", "Sensibilidade, teste de rotacao, bootstrap por episodio, poder preditivo e subamostras."),
    ("", ""),
    ("Fontes", ""),
    ("BPA 12 meses a frente", "Ibovespa Best EPS.xlsx, planilha Sheet1, de 02/01/2006 a 24/08/2026."),
    ("Ibovespa", "BCB SGS serie 7 ate 30/12/2013 e serie B3 do projeto desde 02/01/2014. Emenda verificada, retorno de -2,26% em 02/01/2014."),
    ("Regimes de juros", "swap_pre_di_regimes.csv. Nivel do swap em swap_pre_di_360.csv."),
    ("IPCA", "BCB SGS 433, arquivos JSON em tmp/earnings_regimes_analysis/ipca, de jan/2006 a jul/2026."),
    ("CDI", "BCB SGS 12, diario. De 2006 a 2013 baixado do BCB nesta sessao; de 2014 em diante do arquivo cdi_sgs12.csv ja existente no projeto."),
    ("USDBRL", "PTAX diaria do BCB, arquivos em .codex_eps_analysis/macro."),
    ("", ""),
    ("O que nao foi reproduzido", ""),
    ("SMLL, IMA-B 5 e IMA-B 5+", "Nao ha essas series no projeto. As linhas correspondentes da tabela de classes de ativos ficaram de fora."),
    ("Setores e cestas de fatores", "As abas setoriais MSCI em datasets/ibovespa_msci_return_decomposition sao templates do Bloomberg sem dados carregados. Com um refresh do Bloomberg, as tabelas setoriais e de fatores do report passam a ser reproduziveis."),
    ("Limiar de mudanca de regime", "O report da XP mostra o nivel de swap que dispararia a virada. Isso depende do modelo proprietario de juros e nao foi replicado."),
    ("", ""),
    ("Como o arquivo esta montado", "O Painel_diario e a base de dados. As tabelas-resumo sao formulas com SUMIFS, COUNTIFS e INDEX, entao recalculam sozinhas se a base for estendida. A classificacao de regime em si vem do algoritmo documentado e entra como valor, porque a maquina de estados de confirmacao mensal nao se escreve em formula de planilha."),
    ("Serie de BPA", "A coluna do BPA 12 meses a frente foi removida desta versao publica por ser dado licenciado da Bloomberg (campo BEst EPS). A inclinacao de 63 pregoes, que e a estatistica usada pela regra, continua na planilha, e todas as tabelas e graficos funcionam sem a serie bruta. Para regenerar o arquivo completo, basta ter o Ibovespa Best EPS.xlsx e rodar src/build_report.py.") if PUBLIC else ("", ""),
    ("Aviso de leitura", "Os quatro numeros de retorno sao decomposicao historica, nao previsao. A aba Diagnostico mostra que a ordenacao entre quadrantes e estavel, mas a magnitude do efeito de earnings nao e."),
]
r = 4
for a, b in info:
    if a and not b:
        wl.cell(r, 1, a).font = H2
    elif a:
        wl.cell(r, 1, a).font = Font(name=ARIAL, size=10, bold=True)
        c = wl.cell(r, 2, b)
        c.font = TXT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        wl.row_dimensions[r].height = 30 if len(b) > 110 else 15
    r += 1
wl.cell(r + 1, 1, "Gerado em 02/09/2026. Nenhum arquivo de entrada foi alterado.").font = SMALL

ordem = ["Leia-me", "Resumo_regimes", "Graficos", "Episodios", "Ep_por_quadrante", "Classes_ativos",
         "Transicoes", "Trajetoria_eventos", "Diagnostico", "Painel_diario", "Mensal"]
wb._sheets = [wb[t] for t in ordem]

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False

wb.save(OUT)
print("salvo:", OUT)
